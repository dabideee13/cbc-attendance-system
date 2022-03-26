import re
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


DATA_PATH = Path.joinpath(Path.cwd(), "data")
RAW_PATH = Path.joinpath(DATA_PATH, "raw")
PROCESSED_PATH = Path.joinpath(DATA_PATH, "processed")


def extract_initial(name: str) -> str:
    components = name.split()

    for comp in components:
        if comp.endswith(".") and (comp.lower() != "jr.") and (comp.lower() != "sr."):
            return comp

    return ""


def clean_name(name: str) -> str:
    name = name.replace(extract_initial(name), "")
    return re.sub(r" +", " ", name)


def get_all_files(path: Union[str, Path]) -> list[Path]:
    return list(Path(path).glob("*.xlsx"))


def extract_date(filename: str) -> str:
    pattern = r"\d{1,2}[-./]\d{1,2}[-./]\d{2,4}"
    return re.search(pattern, filename).group()


def format_date(date: str) -> str:
    elem = date.split(".")
    fst, snd, _ = elem

    if len(elem[0]) == 1:
        fst = "0" + elem[0]

    if len(elem[1]) == 1:
        snd = "0" + elem[1]

    return fst + "." + snd + "." + elem[-1]


def format_name(filename: str) -> str:
    name = "Attendance Report"
    return f"{name} - {format_date(extract_date(filename))}" + Path(filename).suffix


def dept_dict(df: pd.DataFrame) -> dict[str, str]:
    return {
        df.loc[index, "Name"]: df.loc[index, "Dept"]
        for index in range(len(df))
    }


def age_dict(df: pd.DataFrame) -> dict[str, str]:
    return {
        df.loc[index, "Name"]: df.loc[index, "Age"]
        for index in range(len(df))
    }


def gender_dict(df: pd.DataFrame) -> dict[str, str]:
    return {
        df.loc[index, "Name"]: df.loc[index, "Gender"]
        for index in range(len(df))
    }


def mapper(name: str, dept_dict: dict[str, str]) -> str:
    try:
        return dept_dict[name]

    except KeyError:
        dept_dict[name] = ""
        return dept_dict[name]


def wrangle(in_file: Union[str, Path], out_file: Union[str, Path]) -> None:

    wb = load_workbook(in_file)
    sheet = wb['Sheet1']

    sheet.delete_cols(idx=2, amount=2)
    sheet.delete_cols(idx=4, amount=8)
    sheet['E2'] = 'Present/Absent'
    sheet['E2'].font = Font(bold=True)

    for row in range(3, sheet.max_row + 1):
        scan = sheet.cell(row, 2)

        if scan.value is not None:
            attendance = sheet.cell(row, 5)
            attendance.value = "Present"
        else:
            attendance = sheet.cell(row, 5)
            attendance.value = "Absent"

    df = pd.DataFrame(sheet.values)
    df = (
        df
        .rename(columns=df.iloc[1])
        .drop(df.index[:2])
        .reset_index()
        .drop("index", axis=1)
    )
    df.Name = df.Name.apply(clean_name)

    file_dept = Path.joinpath(Path.cwd(), "data", "dept.csv")
    df_dept = pd.read_csv(file_dept)
    dept_values = dept_dict(df_dept)
    age_values = age_dict(df_dept)
    gender_values = gender_dict(df_dept)

    df["Department"] = df.Name.apply(lambda x: mapper(x, dept_values))
    df.to_excel(out_file, index=False)

    df["Age"] = df.Name.apply(lambda x: mapper(x, age_values))
    df.to_excel(out_file, index=False)

    df["Gender"] = df.Name.apply(lambda x: mapper(x, gender_values))
    df.to_excel(out_file, index=False)


def main():

    for file in get_all_files(RAW_PATH):
        try:
            wrangle(
                file,
                Path.joinpath(PROCESSED_PATH, format_name(file.name))
            )

        except AttributeError:
            continue


if __name__ == "__main__":
    main()
