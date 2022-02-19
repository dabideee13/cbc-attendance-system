import re
from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.styles import Font

DATA_PATH = Path.joinpath(Path.cwd(), "data")
RAW_PATH = Path.joinpath(DATA_PATH, "raw")
PROCESSED_PATH = Path.joinpath(DATA_PATH, "processed")

SAMPLE = "Attendance Report for Covenant Baptist Church Iligan - Datang (1.16.2022).xlsx"
SAMPLE = "Attendance Report for Covenant Baptist Church Iligan - Datang(12.19.2021).xlsx"

def get_all_files(path: Union[str, Path]) -> list[Path]:
    return list(Path(path).glob("*.xlsx"))


# def extract_date(filename: str) -> str:
#     pattern = r"(?<=\()(.*?)(?=\))"
#     return re.search(pattern, filename).group()

sample2 = "12.5.2022"
sample = "1.12.2021"
sample3 = "1.5.2022"


def extract_date(filename: str) -> str:
    pattern = r"\d{1,2}[-./]\d{1,2}[-./]\d{2,4}"
    return re.search(pattern, filename).group()


def format_date(date: str) -> str:
    elem = date.split(".")
    fst, snd, _ = elem

    if len(elem[0]) == 1:
        fst = "0" + elem[0]

    if len(elem[1]) == 1:
        snd = "0" + elem[1]

    return fst + "." + snd + "." + elem[-1]


def format_name(filename: str) -> str:
    name = "Attendance Report"
    return f"{name} - {format_date(extract_date(filename))}" + Path(filename).suffix


def wrangle(in_file: Union[str, Path], out_file: Union[str, Path]) -> None:

    wb = load_workbook(in_file)
    sheet = wb['Sheet1']

    sheet.delete_cols(idx=2, amount=2)
    sheet.delete_cols(idx=4, amount=8)
    sheet['E2'] = 'Present/Absent'
    sheet['E2'].font = Font(bold=True)

    for row in range(3, sheet.max_row + 1):
        scan = sheet.cell(row, 2)

        if scan.value is not None:
            attendance = sheet.cell(row, 5)
            attendance.value = "Present"
        else:
            attendance = sheet.cell(row, 5)
            attendance.value = "Absent"

    wb.save(out_file)


def main():

    for file in get_all_files(RAW_PATH):
        try:
            wrangle(
                file,
                Path.joinpath(PROCESSED_PATH, format_name(file.name))
            )

        except AttributeError:
            continue


if __name__ == "__main__":
    main()
